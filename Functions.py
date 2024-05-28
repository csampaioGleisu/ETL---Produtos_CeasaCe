from openpyxl import load_workbook, Workbook


def formatar_primeira_coluna(filename):
    wb = load_workbook(filename)
    ws = wb.active
    dados = ['ALFACE', 'ABÓBORA CABOCLO', 'ABÓBORA LEITE', 'CHUCHU', 'FEIJÃO VERDE', 'PIMENTA DE CHEIRO', 'PIMENTÃO', 'REPOLHO', 'TOMATE',
             'ALHO IMPORTADO', 'ALHO NACIONAL', 'BATATA DOCE', 'BATATA INGLÊSA', 'BETERRABA', 'CEBOLA PÊRA NAC.IMP.', 'CENOURA', 'MACAXEIRA', 'MILHO VERDE']
    
    for column in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=1):
        if column[0].row % 2 != 0:
            for cell in column:
                cell_acc = cell.value
                try:
                    if cell_acc not in dados:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value is not None:
                            next_next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                            if next_next_cell.value is not None:
                                next_next_next_cell = ws.cell(row=cell.row, column=cell.column + 4)
                                next_next_next_cell.value = next_next_cell.value
                            next_next_cell.value = next_cell.value
                        next_cell.value = cell_acc
                        cell.value = None
                except Exception as e:
                    print(f"Erro ao processar a célula {cell.coordinate}: {e}")
    
    wb.save(filename)
    
    
def preencher_celulas_em_branco(filename):
    wb = load_workbook(filename)
    ws = wb.active
    
    #
    # Esse laço de repetição linhas (205 a 208), faz com que quando uma célula estiver vazia na 1ª coluna ele pega o valor da célula acima e copia para a célula atual
    #
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value is None:
                cell.value = cell.offset(row=-1).value
                
                
    #
    # Esse laço de repetição faz com que quando uma célula estiver vazia na 4ª coluna ele pega o valor da célula acima e copia para a célula atual
    #
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                if cell.value is None:
                    cell.value = cell.offset(row=-1).value
                    
    #
    # Faz com que o arquivo seja salvo na nova formatação
    #
    wb.save(filename)


def drop_colunm(filename):
    wb = load_workbook(filename)
    ws = wb.active
    
    col_indices_to_remove = []
    for col in ws.iter_cols(1, ws.max_column):
        header = col[0].value
        if header in ["Volume Total ", "(%)"]:
            col_indices_to_remove.append(col[0].column)

    # Exclua as colunas das últimas para as primeiras para evitar mudanças de índice
    for col_index in sorted(col_indices_to_remove, reverse=True):
        ws.delete_cols(col_index)

    wb.save(filename)